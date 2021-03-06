import os

from flask import Flask, flash, render_template, request, jsonify, url_for
from flask_wtf import FlaskForm
from flask_wtf.file import FileRequired, FileAllowed
from openpyxl import load_workbook
from werkzeug.utils import secure_filename, redirect
from wtforms import SelectField, StringField, FileField, RadioField

import db
import graph_test

UPLOAD_FOLDER = './/upload'
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

app = Flask(__name__)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = 'super secret key'
app.config['JSON_SORT_KEYS'] = False


class GenForm(FlaskForm):
    CRN = SelectField('CRN', choices=[])
    semester = SelectField('semester', choices=[])
    exam = SelectField('exam', choices=[])
    type = RadioField('type', choices=[(3, 'Exam'), (1, 'Class'), (2, 'Semester')], default=3)
    color_passive = SelectField('color_passive', choices=[('r', 'Red'), ('g', 'Green'), ('b', 'Blue'),
                                                          ('c', 'Cyan'), ('m', 'Magenta'), ('y', 'Yellow'),
                                                          ('k', 'Black')])

    color_active = SelectField('color_active', choices=[('r', 'Red'), ('g', 'Green'), ('b', 'Blue'),
                                                        ('c', 'Cyan'), ('m', 'Magenta'), ('y', 'Yellow'),
                                                        ('k', 'Black')], default='b')

    upload = FileField('upload', validators=[
        FileRequired(),
        FileAllowed(['xls', 'xlsx'], 'Excel files only!')
    ])


class AddClassForm(FlaskForm):
    CRN = StringField('CRN')
    class_name = StringField('class_name')
    class_num = StringField('class_num')
    semester = StringField('class_num')
    students = StringField('class_num')


class AddExamForm(FlaskForm):
    class_id = SelectField('class_id', choices=[])


class RemoveExamForm(FlaskForm):
    class_id = SelectField('class_id', choices=[])


class RemoveClassForm(FlaskForm):
    class_id = SelectField('class_id', choices=[])


class ViewTables(FlaskForm):
    tables = SelectField('tables', choices=[])


@app.route('/', methods=['GET', 'POST'])
def index():
    form = GenForm()

    conn = db.db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT semester FROM Class")
    semesters = cursor.fetchall()
    conn.close()
    cursor.close()

    form.semester.choices = [("", "---")] + [(semesters[0], semesters[0]) for semesters in semesters]
    form.CRN.choices = [("", "---")]
    form.exam.choices = [("", "---")]

    semester = form.semester.data
    class_id = form.CRN.data
    exam = form.exam.data
    color_passive = form.color_passive.data
    color_active = form.color_active.data
    graph_type = form.type.data

    filename = ''
    if form.validate_on_submit():
        upload = form.upload.data
        filename = secure_filename(upload.filename)
        upload.save(os.path.join(app.instance_path, 'upload', filename))
        return redirect(url_for('index'))

    if request.method == 'POST':
            if graph_type == '1':
                filename = graph_test.test1(class_id)
            elif graph_type == '2':
                filename = graph_test.test2(semester)
            elif graph_type == '3':
                filename = graph_test.test(exam, class_id, color_passive, color_active)
    print(filename)
    return render_template('index.html', form=form, heading="Analyze", filename=filename)


@app.route('/add', methods=['GET', 'POST'])
def add():
    add_class_form = AddClassForm()
    CRN = add_class_form.CRN.data
    class_name = add_class_form.class_name.data
    class_num = add_class_form.class_num.data
    semester = add_class_form.semester.data
    students = add_class_form.students.data

    add_exam_form = GenForm()
    conn = db.db_conn()
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT semester FROM Class")
    semesters = cursor.fetchall()

    add_exam_form.semester.choices = [("", "---")] + [(semesters[0], semesters[0]) for semesters in semesters]
    add_exam_form.CRN.choices = [("", "---")]

    add_exam_CRN = add_exam_form.CRN.data

    add_grades_form = GenForm()

    add_grades_form.semester.choices = [("", "---")] + [(semesters[0], semesters[0]) for semesters in semesters]
    add_grades_form.CRN.choices = [("", "---")]
    add_grades_form.exam.choices = [("", "---")]
    upload_grade = add_grades_form.upload.data

    semester_grade = add_grades_form.semester.data
    class_id_grade = add_grades_form.CRN.data
    exam_grade = add_grades_form.exam.data

    if "add_class" in request.form:
        if CRN is "" or class_name is "" or class_num is "" or semester is "" or students is "":
            flash("Input error", "cat3")
        elif not CRN.isdigit():
            flash("Invalid CRN", "cat3")
        elif not class_num.isdigit():
            flash("Invalid class number", "cat3")
        elif not students.isdigit():
            flash("Invalid student number", "cat3")
        else:
            cursor.execute(
                "INSERT INTO Class (CRN, class_name, class_num, semester)"
                "VALUES('{}', '{}', '{}', '{}')".format(CRN, class_name, class_num, semester))
            conn.commit()
            cursor.execute(
                "SELECT class_id FROM Class WHERE CRN = '{}' AND class_name = '{}' AND class_num = '{}' AND semester "
                "= '{}'".format(
                    CRN, class_name, class_num, semester))
            class_id = cursor.fetchall()
            ids_list = []
            for i in range(int(students)):
                cursor.execute("INSERT INTO Student (class_id) VALUES('{}')".format(class_id[0][0]))
                conn.commit()
                cursor.execute("SELECT student_identifier FROM Student Where class_id = {}".format(class_id[0][0]))
                id_tmp = cursor.fetchall()
                ids_list.append(id_tmp[i][0])
            students_range = str(ids_list[0]) + ' - ' + str(ids_list[len(ids_list) - 1])
            cursor.close()
            conn.close()
            flash("New Class ID: {} with student ids from {}".format(class_id[0][0], students_range), "cat4")

    elif "add_exam" in request.form:
        class_id = add_exam_CRN
        print(class_id)
        if class_id is "":
            flash("Input error", "cat1")
        else:
            cursor.execute("SELECT exam_num FROM Exam WHERE class_id = '{}'".format(
                class_id))
            exam_nums = cursor.fetchall()
            exam_num = 1
            if len(exam_nums) is not 0:
                exam_num = exam_nums[-1][0] + 1
            print(class_id)
            if exam_num > 3:
                flash('This class already has 3 exams', 'cat2')
            else:
                cursor.execute("INSERT INTO Exam (class_id, semester, exam_num) VALUES('{}', (SELECT semester FROM "
                               "Class WHERE class_id = {}), '{}')".format(class_id, class_id, exam_num))
                cursor.execute("SELECT exam_id FROM Exam WHERE class_id = '{}' AND exam_num = '{}'".format(
                    class_id, exam_num))
                exam_id = cursor.fetchall()
                flash("New Exam ID: {}".format(exam_id[0][0]), "cat1")
                conn.commit()
            conn.close()
            cursor.close()
    elif "add_grades" in request.form:
        filename = secure_filename(upload_grade.filename)
        path = os.path.join('grades', filename)
        upload_grade.save(path)
        wb = load_workbook(path)
        ws = wb.active
        max_row = ws.max_row
        ids_tmp = ws['A2:A' + str(max_row)]
        grades_tmp = ws['B2:B' + str(max_row)]
        ids = []
        grades = []

        for i in range(len(ids_tmp)):
            ids.append(ids_tmp[i][0].value)

        for i in range(len(grades_tmp)):
            grades.append(grades_tmp[i][0].value)

        grades_dict = dict(zip(ids, grades))

        cursor.execute("SELECT student_identifier FROM Responses WHERE class_id = {} AND exam_num = {}".format(class_id_grade, exam_grade))
        verify = cursor.fetchall()
        print(verify)

        if verify == verify:
            count = 0
            for student_id, grade in grades_dict.items():
                print(student_id, grade)
                cursor.execute(
                    "SELECT response_id FROM Responses WHERE student_identifier = {} AND class_id = {} AND exam_num = {}".format(
                        student_id, class_id_grade, exam_grade))
                response_id_tmp = cursor.fetchall()
                if response_id_tmp != []:
                    print(response_id_tmp)
                    cursor.execute(
                        "UPDATE Responses SET grade = {} WHERE response_id = {}".format(grade, response_id_tmp[0][0]))
                    count = count+1
            conn.commit()
            flash("{} rows affected".format(count), "cat7")
    conn.close()
    cursor.close()
    return render_template('add.html', add_class_form=add_class_form, add_exam_form=add_exam_form, heading="Edit",
                           add_grades_form=add_grades_form)


@app.route('/view', methods=['GET', 'POST'])
def view():
    tables_select = ViewTables()

    conn = db.db_conn()
    cursor = conn.cursor()
    cursor.execute("SHOW TABLES")
    table_list = cursor.fetchall()
    conn.close()
    cursor.close()

    tables_select.tables.choices = [("", "---")] + [(table_list[0], table_list[0]) for table_list in table_list]

    return render_template('view.html', tables_select=tables_select, heading="View")


@app.route('/remove', methods=['GET', 'POST'])
def remove():
    conn = db.db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT semester FROM Class")
    semesters = cursor.fetchall()

    remove_class_form = GenForm()
    remove_class_form.semester.choices = [("", "---")] + [(semesters[0], semesters[0]) for semesters in semesters]
    remove_class_form.CRN.choices = [("", "---")]

    remove_class_CRN = remove_class_form.CRN.data

    remove_exam_form = GenForm()
    remove_exam_form.semester.choices = [("", "---")] + [(semesters[0], semesters[0]) for semesters in semesters]
    remove_exam_form.CRN.choices = [("", "---")]
    remove_exam_form.exam.choices = [("", "---")]

    remove_exam_CRN = remove_exam_form.CRN.data
    remove_exam_exam = remove_exam_form.exam.data

    if "remove_class" in request.form:
        remove_class = remove_class_CRN
        cursor.execute("SELECT * FROM Class WHERE class_id = {}".format(remove_class))
        confirm = cursor.fetchall()[0]
        cursor.execute("DELETE FROM Responses Where class_id = {}".format(remove_class))
        cursor.execute("DELETE FROM methods_used Where class_id = {}".format(remove_class))
        cursor.execute("DELETE FROM Exam Where class_id = {}".format(remove_class))
        cursor.execute("DELETE FROM Student WHERE Class_id = {}".format(remove_class))
        cursor.execute("DELETE FROM Class Where class_id = {}".format(remove_class))
        flash("Removed class {} {} {} {} {}".format(confirm[0], confirm[1], confirm[2], confirm[3], confirm[4]), 'cat5')
        conn.commit()

    elif "remove_exam" in request.form:
        cursor.execute("SELECT exam_id FROM Exam WHERE class_id = {} AND exam_num = {}".format(remove_exam_CRN, remove_exam_exam))
        remove_exam = cursor.fetchall()[0][0]
        cursor.execute("SELECT * FROM Class WHERE class_id = {}".format(remove_exam_CRN))
        confirm = cursor.fetchall()[0]
        print(remove_exam)
        cursor.execute("DELETE FROM Responses Where exam_num = {} AND class_id = {}".format(remove_exam_exam, remove_exam_CRN))
        cursor.execute("DELETE FROM methods_used Where exam_id = {} AND class_id = {}".format(remove_exam_exam, remove_exam_CRN))
        cursor.execute("DELETE FROM Exam Where exam_id = {}".format(remove_exam))
        flash("Removed exam from {} {} {} {} {}".format(confirm[0], confirm[1], confirm[2], confirm[3], confirm[4]), 'cat6')
        conn.commit()
    cursor.close()
    conn.close()
    return render_template('remove.html', remove_class_form=remove_class_form, remove_exam_form=remove_exam_form,
                           heading="Remove")


@app.route('/CRN/<semester>')
def get_crn(semester):
    conn = db.db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT class_id, CRN FROM Class Where semester = '{}';".format(semester))
    CRNs = cursor.fetchall()
    conn.close()
    cursor.close()
    CRN_Array = []

    for CRN in CRNs:
        CRN_Obj = {'class_id': CRN[0], 'CRN': CRN[1]}
        CRN_Array.append(CRN_Obj)

    return jsonify({'CRNs': CRN_Array})


@app.route('/test/<class_id>')
def get_test(class_id):
    conn = db.db_conn()
    cursor = conn.cursor()
    cursor.execute(
        "select DISTINCT exam_id, exam_num from Exam INNER JOIN Class ON Class.class_id = Exam.class_id WHERE "
        "Class.class_id = {}".format(class_id))
    tests = cursor.fetchall()
    print(tests)
    conn.close()
    cursor.close()
    testsArray = []

    for test in tests:
        testObj = {'exam_id': test[0], 'exam_num': test[1]}
        testsArray.append(testObj)

    return jsonify({'tests': testsArray})


@app.route('/table/<table>')
def get_table(table):
    conn = db.db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM StudyStrategies1.{}".format(table))
    table_vals = cursor.fetchall()
    cursor.execute("SELECT COLUMN_NAME FROM information_schema.columns WHERE TABLE_SCHEMA = 'StudyStrategies1' AND "
                   "TABLE_NAME = '{}';".format(table))
    table_cols = cursor.fetchall()
    print(table_cols)
    conn.close()
    cursor.close()
    test_array = []

    for i in range(len(table_vals)):
        testObj = {}
        for j in range(len(table_cols)):
            testObj.update({table_cols[j][0]: table_vals[i][j]})
        test_array.append(testObj)

    return jsonify(test_array)


@app.errorhandler(404)
def page_not_found(e):
    # note that we set the 404 status explicitly
    return render_template('404.html'), 404
